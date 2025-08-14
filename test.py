import speech_recognition as sr
import tkinter as tk
from tkinter import ttk, filedialog
import threading
import queue
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class SpeechToTextApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Real-Time Speech to Text Converter")
        self.root.geometry("800x600")

        # Recognition variables
        self.recognizer = sr.Recognizer()
        self.is_listening = False
        self.audio_queue = queue.Queue()
        self.current_language = "en-US"
        
        # Excel file path
        self.excel_path = r"D:\Scan It\Voice\product order.xlsx"
        self.initialize_excel()

        # Create GUI
        self.create_widgets()

        # Configure microphone
        self.microphone = sr.Microphone()
        with self.microphone as source:
            self.recognizer.adjust_for_ambient_noise(source, duration=1)

    def initialize_excel(self):
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
        
        # Create or load Excel file
        if not os.path.exists(self.excel_path):
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Voice Orders"
            # Add headers if new file
            self.ws.append(["Timestamp", "Order Text"])
            self.wb.save(self.excel_path)
        else:
            self.wb = load_workbook(self.excel_path)
            self.ws = self.wb.active

    def create_widgets(self):
        # Set dark theme colors
        bg_color = "#181818"
        fg_color = "#F5F5F5"
        entry_bg = "#222222"
        accent = "#333333"

        self.root.configure(bg=bg_color)

        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.configure(style="Dark.TFrame")

        # Controls frame
        controls_frame = ttk.Frame(main_frame)
        controls_frame.pack(fill=tk.X, pady=5)
        controls_frame.configure(style="Dark.TFrame")

        # Set ttk styles for dark mode
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Dark.TFrame", background=bg_color)
        style.configure("Dark.TLabel", background=bg_color, foreground=fg_color)
        style.configure("Dark.TButton", background=accent, foreground=fg_color)
        style.configure("Dark.TCombobox", fieldbackground=entry_bg, background=entry_bg, foreground=fg_color)
        style.map("Dark.TButton", background=[("active", accent)])

        # Language selection
        ttk.Label(controls_frame, text="Language:", style="Dark.TLabel").pack(side=tk.LEFT)
        self.language_var = tk.StringVar(value="en-US")
        languages = {
            "English (US)": "en-US",
            "English (UK)": "en-GB",
            "Spanish": "es-ES",
            "French": "fr-FR",
            "German": "de-DE",
            "Italian": "it-IT",
            "Bengali": "bn-BD"
        }
        self.language_menu = ttk.Combobox(
            controls_frame,
            textvariable=self.language_var,
            values=list(languages.keys()),
            state="readonly",
            width=15
        )
        self.language_menu.pack(side=tk.LEFT, padx=5)
        self.language_menu.bind("<<ComboboxSelected>>", self.update_language)

        # Engine selection
        ttk.Label(controls_frame, text="Engine:", style="Dark.TLabel").pack(side=tk.LEFT, padx=(10,0))
        self.engine_var = tk.StringVar(value="Google")
        engines = ["Google", "Sphinx (offline)"]
        self.engine_menu = ttk.Combobox(
            controls_frame,
            textvariable=self.engine_var,
            values=engines,
            state="readonly",
            width=15
        )
        self.engine_menu.pack(side=tk.LEFT, padx=5)

        # Buttons
        self.listen_button = ttk.Button(
            controls_frame,
            text="Start Listening",
            command=self.toggle_listening,
            style="Dark.TButton"
        )
        self.listen_button.pack(side=tk.LEFT, padx=10)

        self.save_button = ttk.Button(
            controls_frame,
            text="Save Transcript",
            command=self.save_transcript,
            state=tk.DISABLED,
            style="Dark.TButton"
        )
        self.save_button.pack(side=tk.LEFT)

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        self.status_bar = ttk.Label(
            main_frame,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor=tk.W,
            style="Dark.TLabel"
        )
        self.status_bar.pack(fill=tk.X, pady=(5,0))

        # Text display (use tk.Text for better color control)
        self.text_display = tk.Text(
            main_frame,
            wrap=tk.WORD,
            font=('Arial', 12),
            padx=10,
            pady=10,
            bg=entry_bg,
            fg=fg_color,
            insertbackground=fg_color,
            selectbackground="#444444"
        )
        self.text_display.pack(fill=tk.BOTH, expand=True)

        # Add timestamp to first line
        self.add_timestamp()

    def update_language(self, event=None):
        language_map = {
            "English (US)": "en-US",
            "English (UK)": "en-GB",
            "Spanish": "es-ES",
            "French": "fr-FR",
            "German": "de-DE",
            "Italian": "it-IT",
            "Bengali": "bn-BD"
        }
        self.current_language = language_map.get(self.language_var.get(), "en-US")

    def toggle_listening(self):
        if self.is_listening:
            self.stop_listening()
        else:
            self.start_listening()

    def start_listening(self):
        self.is_listening = True
        self.listen_button.config(text="Stop Listening")
        self.status_var.set("Listening...")

        # Clear any previous audio in the queue
        while not self.audio_queue.empty():
            self.audio_queue.get()

        # Start background thread for listening
        self.listening_thread = threading.Thread(
            target=self.listen_loop,
            daemon=True
        )
        self.listening_thread.start()

        # Start processing thread
        self.processing_thread = threading.Thread(
            target=self.process_audio_queue,
            daemon=True
        )
        self.processing_thread.start()

    def stop_listening(self):
        self.is_listening = False
        self.listen_button.config(text="Start Listening")
        self.status_var.set("Ready")
        self.save_button.config(state=tk.NORMAL)

    def listen_loop(self):
        with self.microphone as source:
            while self.is_listening:
                try:
                    audio = self.recognizer.listen(
                        source,
                        timeout=2,
                        phrase_time_limit=5
                    )
                    self.audio_queue.put(audio)
                except sr.WaitTimeoutError:
                    continue
                except Exception as e:
                    self.status_var.set(f"Error: {str(e)}")
                    self.stop_listening()
                    break

    def process_audio_queue(self):
        while self.is_listening or not self.audio_queue.empty():
            try:
                audio = self.audio_queue.get(timeout=1)
                self.process_audio(audio)
            except queue.Empty:
                continue

    def process_audio(self, audio):
        try:
            engine = self.engine_var.get()

            if engine == "Google":
                text = self.recognizer.recognize_google(
                    audio,
                    language=self.current_language
                )
            elif engine == "Sphinx (offline)":
                text = self.recognizer.recognize_sphinx(audio)
            else:
                text = "Unsupported engine"

            self.display_text(text)
            self.save_to_excel(text)

        except sr.UnknownValueError:
            self.status_var.set("Could not understand audio")
        except sr.RequestError as e:
            self.status_var.set(f"API error: {str(e)}")
        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")

    def save_to_excel(self, text):
        try:
            # Get current timestamp
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Append to Excel file
            self.ws.append([timestamp, text])
            self.wb.save(self.excel_path)
            
            self.status_var.set(f"Text saved to Excel: {os.path.basename(self.excel_path)}")
        except Exception as e:
            self.status_var.set(f"Error saving to Excel: {str(e)}")

    def display_text(self, text):
        self.text_display.insert(tk.END, text + "\n")
        self.text_display.see(tk.END)
        self.root.update()

    def add_timestamp(self):
        now = datetime.datetime.now()
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
        self.text_display.insert(tk.END, f"=== Session started: {timestamp} ===\n\n")

    def save_transcript(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
            initialfile=f"transcript_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )

        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(self.text_display.get("1.0", tk.END))
                self.status_var.set(f"Transcript saved to {os.path.basename(file_path)}")
            except Exception as e:
                self.status_var.set(f"Error saving file: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SpeechToTextApp(root)
    root.mainloop()