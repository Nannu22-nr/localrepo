import gradio as gr
import speech_recognition as sr
import tempfile
import datetime
import whisper
import os
from openpyxl import Workbook, load_workbook

# Setup Excel path
excel_path = r"D:\Scan It\Voice\product order.xlsx"
# Supported languages
language_options = {
    "English (US)": "en-US",
    "English (UK)": "en-GB",
    "Spanish": "es-ES",
    "French": "fr-FR",
    "German": "de-DE",
    "Italian": "it-IT",
    "Bengali": "bn-BD"
}

# Initialize Excel workbook
def initialize_excel():
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Transcripts"
        ws.append(["Timestamp", "Text"])
        wb.save(excel_path)
    else:
        wb = load_workbook(excel_path)
    return wb

# Save transcript to Excel
def save_to_excel(text):
    wb = initialize_excel()
    ws = wb.active
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, text])
    wb.save(excel_path)

# Transcribe from microphone
def transcribe_audio(engine, language_label):
    language = language_options.get(language_label, "en-US")
    recognizer = sr.Recognizer()
    try:
        with sr.Microphone() as source:
            recognizer.adjust_for_ambient_noise(source, duration=1)
            audio = recognizer.listen(source, timeout=5, phrase_time_limit=7)

        if engine == "Google":
            text = recognizer.recognize_google(audio, language=language)

        elif engine == "Whisper (local)":
            model = whisper.load_model("base")
            with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as f:
                f.write(audio.get_wav_data())
                f.flush()
                result = model.transcribe(f.name)
                text = result["text"]
            os.remove(f.name)

        elif engine == "Sphinx (offline)":
            text = recognizer.recognize_sphinx(audio)
        else:
            text = "Unsupported engine selected."

        save_to_excel(text)
        return f"📝 {text}\n\n✅ Saved at {datetime.datetime.now().strftime('%H:%M:%S')}"

    except sr.UnknownValueError:
        return "🤷 Speech not understood."
    except sr.RequestError as e:
        return f"🚨 API error: {str(e)}"
    except Exception as e:
        return f"⚠️ Error: {str(e)}"

# Gradio Interface
with gr.Blocks(theme="soft") as demo:
    gr.Markdown("## 🎙️ Real-Time Speech to Text Converter")
    gr.Markdown("Select language and engine, then click **Start Listening** to transcribe your voice.")

    with gr.Row():
        lang_dropdown = gr.Dropdown(label="Language", choices=list(language_options.keys()), value="English (US)")
        engine_dropdown = gr.Dropdown(label="Engine", choices=["Google", "Whisper (local)", "Sphinx (offline)"], value="Google")

    output_textbox = gr.Textbox(label="Transcript", lines=10)
    listen_button = gr.Button("🎧 Start Listening")

    listen_button.click(
        fn=transcribe_audio,
        inputs=[engine_dropdown, lang_dropdown],
        outputs=output_textbox
    )

demo.launch()
